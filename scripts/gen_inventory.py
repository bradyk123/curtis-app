#!/usr/bin/env python3
"""Generate src/data/inventory.ts + download media from the Track & Field xlsx.

Circuits sheet -> categories & circuit metadata (authoritative circuit list).
Items sheet    -> exercises attached to circuits by name, in row order.
"""
import openpyxl, re, json, os, sys, hashlib

XLSX = "/Users/bradykirtland/Desktop/Track & Field App - Sheets.xlsx"
REPO = "/Users/bradykirtland/Desktop/Curtis APP"
MEDIA_DIR = os.path.join(REPO, "public", "media")

# Orphan circuits referenced by Items but missing from Circuits sheet ->
# best-fit category (flagged to user for confirmation).
ORPHAN_CATEGORY = {
    "Hurdle Series": "Hurdle Mobility",
    "Bow": "Multiple Jump",
    "Devil": "Sprint Drills",
}

def slug(s):
    s = str(s).strip().lower()
    s = s.replace("&", "and").replace("’", "").replace("'", "")
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "item"

wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)

# ---------- Circuits ----------
cs = list(wb["Circuits"].iter_rows(values_only=True))
# cols: 0 RowID,1 Circuit,2 Image,3 Parent,4 Category,5 Reps,6 Rest,7 Notes,8 Notes2
circuit_meta = {}          # name -> dict
category_order = []        # preserve appearance order
circuit_order_in_cat = {}  # category -> [circuit names]
for r in cs[1:]:
    name = r[1]
    if not name:
        continue
    name = str(name).strip()
    cat = str(r[4]).strip() if r[4] else "Uncategorized"
    reps = str(r[5]).strip() if len(r) > 5 and r[5] else None
    rest = str(r[6]).strip() if len(r) > 6 and r[6] else None
    notes = str(r[7]).strip() if len(r) > 7 and r[7] else None
    circuit_meta[name] = {"category": cat, "reps": reps, "rest": rest, "notes": notes}
    if cat not in category_order:
        category_order.append(cat)
        circuit_order_in_cat[cat] = []
    circuit_order_in_cat[cat].append(name)

# ---------- Items ----------
it = list(wb["Items"].iter_rows(values_only=True))
# cols: 0 Item,1 Image,2 Instructions,3 Details, 4..(n-2) Circuit N, last Primary Cue
exercises_by_circuit = {}  # circuit name -> [exercise dict] in order
media_urls = {}            # url -> filename
for r in it[1:]:
    name = r[0]
    if not name:
        continue
    name = str(name).strip()
    img = r[1] if len(r) > 1 else None
    instr = r[2] if len(r) > 2 else None
    details = r[3] if len(r) > 3 else None
    # circuit membership: first non-empty value among the Circuit columns (idx 4..last-1)
    circ = None
    for v in r[4:-1]:
        if v is not None and str(v).strip() != "":
            circ = str(v).strip()
            break
    if circ is None:
        # fall back to last col scan excluding primary cue already handled
        continue
    cue = None
    if instr and str(instr).strip():
        cue = str(instr).strip()
    elif details and str(details).strip():
        cue = str(details).strip()
    media = None
    if img and str(img).strip().lower().endswith(".gif"):
        url = str(img).strip()
        fn = url.split("/")[-1].split("?")[0]
        media_urls[url] = fn
        media = fn
    exercises_by_circuit.setdefault(circ, []).append(
        {"name": name, "media": media, "cues": cue}
    )

# ---------- Merge orphan circuits ----------
for oc, cat in ORPHAN_CATEGORY.items():
    if oc in exercises_by_circuit and oc not in circuit_meta:
        circuit_meta[oc] = {"category": cat, "reps": None, "rest": None, "notes": None}
        if cat not in category_order:
            category_order.append(cat)
            circuit_order_in_cat[cat] = []
        circuit_order_in_cat[cat].append(oc)

# ---------- Build nested structure ----------
def subtitle_for(meta):
    parts = []
    if meta["reps"]:
        parts.append(meta["reps"])
    if meta["notes"] and not meta["reps"]:
        # spacing-style note used as subtitle
        parts.append(meta["notes"])
    if meta["rest"]:
        parts.append(meta["rest"])
    return " / ".join(parts) if parts else None

categories = []
used_ex_ids = set()
for cat in category_order:
    cat_obj = {"id": slug(cat), "name": cat, "circuits": []}
    for cname in circuit_order_in_cat[cat]:
        meta = circuit_meta[cname]
        exs = exercises_by_circuit.get(cname, [])
        # circuits with no items but a multi-line Notes field -> derive exercise list
        if not exs and meta["notes"] and "\n" in meta["notes"]:
            exs = [{"name": ln.strip(), "media": None, "cues": None}
                   for ln in meta["notes"].split("\n") if ln.strip()]
        circuit = {
            "id": slug(cname),
            "name": cname,
            "subtitle": subtitle_for(meta),
            "exercises": [],
        }
        seen = {}
        for e in exs:
            base = slug(e["name"])
            seen[base] = seen.get(base, 0) + 1
            eid = base if seen[base] == 1 else f"{base}-{seen[base]}"
            circuit["exercises"].append(
                {"id": eid, "name": e["name"], "media": e["media"], "cues": e["cues"]}
            )
        cat_obj["circuits"].append(circuit)
    categories.append(cat_obj)

# ---------- Emit inventory.ts ----------
def esc(s):
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n")

lines = []
lines.append('import type { Category } from "../types";')
lines.append("")
lines.append("/**")
lines.append(" * Generated from \"Track & Field App - Sheets.xlsx\" (Circuits + Items sheets).")
lines.append(" * 40 circuits / 247 exercises across 12 categories; GIF demos bundled in")
lines.append(" * public/media and referenced base-aware via MEDIA below. Regenerate with")
lines.append(" * scripts/gen_inventory.py rather than hand-editing.")
lines.append(" */")
lines.append('const MEDIA = import.meta.env.BASE_URL + "media/";')
lines.append("")
lines.append("export const inventory: Category[] = [")
for cat in categories:
    lines.append("  {")
    lines.append(f'    id: "{cat["id"]}",')
    lines.append(f'    name: "{esc(cat["name"])}",')
    lines.append("    circuits: [")
    for c in cat["circuits"]:
        lines.append("      {")
        lines.append(f'        id: "{c["id"]}",')
        lines.append(f'        name: "{esc(c["name"])}",')
        if c["subtitle"]:
            lines.append(f'        subtitle: "{esc(c["subtitle"])}",')
        if not c["exercises"]:
            lines.append("        exercises: [],")
        else:
            lines.append("        exercises: [")
            for e in c["exercises"]:
                parts = [f'id: "{e["id"]}"', f'name: "{esc(e["name"])}"']
                if e["media"]:
                    parts.append(f'mediaUrl: MEDIA + "{e["media"]}"')
                if e["cues"]:
                    parts.append(f'cues: "{esc(e["cues"])}"')
                lines.append("          { " + ", ".join(parts) + " },")
            lines.append("        ],")
        lines.append("      },")
    lines.append("    ],")
    lines.append("  },")
lines.append("];")
lines.append("")

out_ts = os.path.join(REPO, "src", "data", "inventory.ts")
with open(out_ts, "w") as f:
    f.write("\n".join(lines))

# Also emit JSON (consumed by scripts/seed_supabase.mjs to seed the DB).
out_json = os.path.join(REPO, "scripts", "inventory.json")
with open(out_json, "w") as f:
    json.dump(categories, f, indent=2, ensure_ascii=False)
print("wrote inventory.json ->", out_json)

# stats
n_cat = len(categories)
n_circ = sum(len(c["circuits"]) for c in categories)
n_ex = sum(len(ci["exercises"]) for c in categories for ci in c["circuits"])
n_media = sum(1 for c in categories for ci in c["circuits"] for e in ci["exercises"] if e["media"])
print(f"categories={n_cat} circuits={n_circ} exercises={n_ex} with_media={n_media}")
print(f"unique media files to download: {len(media_urls)}")

# save media manifest for the downloader
with open("/private/tmp/claude-501/-Users-bradykirtland-Desktop-Curtis-APP/bc4c2f26-fbcd-4897-81e0-76a902cebf40/scratchpad/media_manifest.json", "w") as f:
    json.dump(media_urls, f, indent=2)
print("wrote inventory.ts ->", out_ts)
